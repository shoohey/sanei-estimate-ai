"""入力済みExcelを正解データ(SurveyData)として読み出すパーサー。

PDF読み取り精度評価のため、既に手入力された現調シート(.xlsx)を
SurveyData形式へマッピングする。セル位置は固定せず、ラベル文字列
("案件名："、"モジュールメーカー" など)をキーに右隣セル/同行セルを
取得する方式で、Excel配置の揺れに対応する。

公開関数:
    load_ground_truth(xlsx_path) -> SurveyData
    discover_ground_truth_dir(base_dir) -> list[(case_name, xlsx_path)]
    find_pdf_for_xlsx(xlsx_path) -> list[str]

CLI:
    python3 tests/ground_truth.py
"""

from __future__ import annotations

import re
import sys
import unicodedata
from pathlib import Path
from typing import Optional

import openpyxl

# プロジェクトルートを sys.path に通して models をインポート可能にする
PROJECT_ROOT = Path(__file__).parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from models.survey_data import (  # noqa: E402
    BTPlacement,
    CInstallation,
    DesignStatus,
    FinalConfirmation,
    GroundType,
    HighVoltageChecklist,
    LocationType,
    PlannedEquipment,
    ProjectInfo,
    SupplementarySheet,
    SurveyData,
)


# ----------------------------------------------------------------------
# ヘルパー
# ----------------------------------------------------------------------

def _normalize(text: object) -> str:
    """全角/半角・空白の正規化。比較用の文字列を生成。"""
    if text is None:
        return ""
    s = str(text)
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("　", " ").replace(" ", " ").replace("\xa0", " ")
    return s.strip()


def _to_float(text: object) -> float:
    """'400W' → 400.0, '150.4KW' → 150.4 など、単位除去して数値化。失敗時は0."""
    s = _normalize(text)
    if not s:
        return 0.0
    # 数字・小数点・マイナスのみ抽出
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return 0.0
    try:
        return float(m.group(0))
    except ValueError:
        return 0.0


def _to_int(text: object) -> int:
    """'376枚' → 376 など。失敗時は0."""
    return int(_to_float(text))


def _strip_label(text: str, label: str) -> str:
    """'案件名：xxx' から label を取り除いて値部分を返す。"""
    s = _normalize(text)
    label_n = _normalize(label)
    # 「案件名：」「案件名:」「案件名 」など
    for sep in ["：", ":", " ", ""]:
        prefix = f"{label_n}{sep}"
        if s.startswith(prefix):
            return s[len(prefix):].strip()
    return s


def _build_cell_index(ws) -> dict[str, list]:
    """シート全体を走査し、テキスト → セルリスト の辞書を作る。"""
    idx: dict[str, list] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            key = _normalize(cell.value)
            if not key:
                continue
            idx.setdefault(key, []).append(cell)
    return idx


def _find_label_cell(ws, *labels: str):
    """labels のいずれかと前方一致するセルを探す。最初に見つかったものを返す。"""
    targets = [_normalize(l) for l in labels]
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            v = _normalize(cell.value)
            for t in targets:
                if v.startswith(t) or t in v:
                    return cell, v
    return None, ""


def _value_after_label(ws, *labels: str) -> str:
    """ラベルセルを探し、その右隣のセル値、なければラベルセル内の値部分を返す。"""
    cell, v = _find_label_cell(ws, *labels)
    if cell is None:
        return ""
    # ラベルが「案件名：xxx」のように同セルに値も入っているケース
    for label in labels:
        stripped = _strip_label(v, label)
        if stripped and stripped != v:
            return stripped
    # 右隣セル(C→E など空セルをスキップしながら2セルまで)
    row = cell.row
    col = cell.column
    for offset in range(1, 5):
        try:
            right = ws.cell(row=row, column=col + offset)
        except Exception:
            break
        if right.value is not None and _normalize(right.value):
            return _normalize(right.value)
    return ""


def _value_for_check(ws, *labels: str) -> str:
    """高圧/別紙チェック項目の「確認内容」列の値を取得。

    通常はラベルセルから2列右(C列ラベル → E列値)。
    """
    cell, _ = _find_label_cell(ws, *labels)
    if cell is None:
        return ""
    # E列(列番号5)を最優先
    target = ws.cell(row=cell.row, column=5)
    if target.value is not None and _normalize(target.value):
        return _normalize(target.value)
    # フォールバック: 右隣を順に探索
    for offset in range(1, 6):
        right = ws.cell(row=cell.row, column=cell.column + offset)
        if right.value is not None and _normalize(right.value):
            return _normalize(right.value)
    return ""


def _to_bool_present(text: str) -> bool:
    """'あり' → True, 'なし' → False。"""
    s = _normalize(text)
    if not s:
        return False
    if "あり" in s:
        return True
    if "なし" in s:
        return False
    # チェックマーク類
    if s in ("○", "◯", "✔", "✓", "有", "yes", "Yes", "YES"):
        return True
    if s in ("×", "✕", "無", "no", "No", "NO"):
        return False
    return False


def _parse_design_status(text: str) -> DesignStatus:
    s = _normalize(text)
    if "確定" in s:
        return DesignStatus.CONFIRMED
    if "仮" in s:
        return DesignStatus.TENTATIVE
    return DesignStatus.UNDECIDED


def _parse_ground_type(text: str) -> GroundType:
    s = _normalize(text).upper()
    # 「A D」「A,D」など複数記載がある場合は最初の文字を採用
    for ch in ("A", "C", "D"):
        if ch in s:
            return GroundType(ch)
    return GroundType.A


def _parse_c_installation(text: str) -> CInstallation:
    s = _normalize(text)
    if "可" in s and "不可" not in s:
        return CInstallation.POSSIBLE
    if "不可" in s:
        return CInstallation.IMPOSSIBLE
    return CInstallation.POSSIBLE


def _parse_location(text: str) -> Optional[LocationType]:
    s = _normalize(text)
    if "屋内" in s:
        return LocationType.INDOOR
    if "屋外" in s:
        return LocationType.OUTDOOR
    return None


def _parse_bt_placement(text: str) -> Optional[BTPlacement]:
    s = _normalize(text)
    if "屋内" in s:
        return BTPlacement.INDOOR
    if "屋外" in s:
        return BTPlacement.OUTDOOR
    if "なし" in s or "設置なし" in s:
        return BTPlacement.NONE
    return None


# ----------------------------------------------------------------------
# 公開関数
# ----------------------------------------------------------------------

def load_ground_truth(xlsx_path: str) -> SurveyData:
    """入力済みExcelからSurveyDataを構築する。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_names = wb.sheetnames

    # メインシート(高圧)
    main_sheet = None
    supp_sheet = None
    for name in sheet_names:
        if "高圧" in name or "★" in name:
            main_sheet = wb[name]
        elif "別紙" in name:
            supp_sheet = wb[name]
    if main_sheet is None:
        main_sheet = wb[sheet_names[0]]

    # ----------- ProjectInfo -----------
    project = ProjectInfo(
        project_name=_value_after_label(main_sheet, "案件名"),
        address=_value_after_label(main_sheet, "所在地"),
        survey_date=_value_after_label(main_sheet, "調査日"),
        weather=_value_after_label(main_sheet, "天気"),
        surveyor=_value_after_label(main_sheet, "調査者"),
    )

    # ----------- PlannedEquipment -----------
    equipment = PlannedEquipment(
        module_maker=_value_for_check(main_sheet, "モジュールメーカー"),
        module_model=_value_for_check(main_sheet, "モジュール型式"),
        module_output_w=_to_float(_value_for_check(main_sheet, "モジュール定格出力")),
        planned_panels=_to_int(_value_for_check(main_sheet, "設置予定枚数")),
        pv_capacity_kw=_to_float(_value_for_check(main_sheet, "想定PV容量")),
        design_status=_parse_design_status(_value_for_check(main_sheet, "設計確定度")),
    )

    # ----------- HighVoltageChecklist -----------
    hv = HighVoltageChecklist(
        building_drawing=_to_bool_present(_value_for_check(main_sheet, "建物図面")),
        single_line_diagram=_to_bool_present(_value_for_check(main_sheet, "単線結線図")),
        ground_type=_parse_ground_type(_value_for_check(main_sheet, "接地種類")),
        c_installation=_parse_c_installation(_value_for_check(main_sheet, "C種別設置可否")),
        vt_available=_to_bool_present(_value_for_check(main_sheet, "VT有無")),
        ct_available=_to_bool_present(_value_for_check(main_sheet, "CT有無")),
        relay_space=_to_bool_present(_value_for_check(main_sheet, "継電器スペース")),
        pcs_space=_to_bool_present(_value_for_check(main_sheet, "PCS設置スペース")),
        pcs_location=_parse_location(_value_for_check(main_sheet, "（ありの場合）")),
        bt_space=_parse_bt_placement(_value_for_check(main_sheet, "BT設置スペース")),
        bt_backup_capacity=_value_for_check(main_sheet, "BTバックアップ回路容量"),
        tr_capacity=_value_for_check(main_sheet, "Tr容量余裕"),
        pre_use_self_check=_to_bool_present(_value_for_check(main_sheet, "使用前自己確認")),
        separation_ns_mm=_to_float(_value_for_check(main_sheet, "離隔 縦")),
        separation_ew_mm=_to_float(_value_for_check(main_sheet, "離隔 横")),
    )

    # ----------- SupplementarySheet -----------
    if supp_sheet is not None:
        supplementary = SupplementarySheet(
            crane_available=_to_bool_present(_value_for_check(supp_sheet, "クレーン")),
            scaffold_location=_value_for_check(supp_sheet, "足場設置予定位置"),
            pole_number=_value_for_check(supp_sheet, "電柱番号"),
            wiring_route=_value_for_check(supp_sheet, "配管、配線ルート"),
            cubicle_location=_to_bool_present(_value_for_check(supp_sheet, "キュービクル")),
            bt_location=_value_for_check(supp_sheet, "BT設置位置"),
            meter_photo=_value_for_check(supp_sheet, "引込柱"),
        )
    else:
        supplementary = SupplementarySheet()

    # ----------- FinalConfirmation -----------
    confirmation = FinalConfirmation(
        surveyor_name=_value_after_label(main_sheet, "調査者（現調実施）"),
        design_reviewer=_value_after_label(main_sheet, "設計確認"),
        works_reviewer=_value_after_label(main_sheet, "ワークス部確認"),
    )

    return SurveyData(
        project=project,
        equipment=equipment,
        high_voltage=hv,
        supplementary=supplementary,
        confirmation=confirmation,
    )


def discover_ground_truth_dir(
    base_dir: str = "見積AI入力資料/入力済み",
) -> list[tuple[str, str]]:
    """入力済みディレクトリ配下の (案件名, xlsxパス) リストを返す。"""
    base = Path(base_dir)
    if not base.is_absolute():
        base = PROJECT_ROOT / base_dir
    if not base.exists():
        return []

    results: list[tuple[str, str]] = []
    for case_dir in sorted(base.iterdir()):
        if not case_dir.is_dir():
            continue
        # 現調シート.xlsx を探す
        xlsx_files = sorted(case_dir.glob("*現調シート*.xlsx"))
        # 一時ファイル(~$..)を除外
        xlsx_files = [p for p in xlsx_files if not p.name.startswith("~$")]
        if not xlsx_files:
            xlsx_files = [
                p for p in sorted(case_dir.glob("*.xlsx"))
                if not p.name.startswith("~$")
            ]
        if xlsx_files:
            results.append((case_dir.name, str(xlsx_files[0])))
    return results


def find_pdf_for_xlsx(xlsx_path: str) -> list[str]:
    """xlsxと同じディレクトリの現調シート関連PDF(メイン+別紙)を返す。"""
    p = Path(xlsx_path)
    case_dir = p.parent
    pdfs: list[str] = []
    # メイン現調シートPDF
    main_candidates = sorted(case_dir.glob("*現調シート.pdf"))
    main_candidates = [
        x for x in main_candidates
        if "別紙" not in x.name and not x.name.startswith("~$")
    ]
    pdfs.extend(str(x) for x in main_candidates)
    # 別紙PDF
    supp_candidates = sorted(case_dir.glob("*現調シート別紙*.pdf"))
    pdfs.extend(str(x) for x in supp_candidates)
    return pdfs


# ----------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------

def _print_case(case_name: str, sd: SurveyData) -> None:
    print(f"\n=== {case_name} ===")
    print(f"  案件名         : {sd.project.project_name}")
    print(f"  所在地         : {sd.project.address}")
    print(f"  調査日         : {sd.project.survey_date}")
    print(f"  モジュールメーカー: {sd.equipment.module_maker}")
    print(f"  モジュール型式 : {sd.equipment.module_model}")
    print(f"  定格出力(W)    : {sd.equipment.module_output_w}")
    print(f"  設置枚数       : {sd.equipment.planned_panels}")
    print(f"  PV容量(kW)     : {sd.equipment.pv_capacity_kw}")
    print(f"  設計確定度     : {sd.equipment.design_status.value}")


if __name__ == "__main__":
    cases = discover_ground_truth_dir()
    print(f"発見した入力済みケース: {len(cases)}件")
    for name, xlsx in cases:
        try:
            sd = load_ground_truth(xlsx)
            _print_case(name, sd)
            pdfs = find_pdf_for_xlsx(xlsx)
            print(f"  関連PDF        : {len(pdfs)}件")
            for pdf in pdfs:
                print(f"    - {Path(pdf).name}")
        except Exception as e:
            print(f"\n=== {name} ===")
            print(f"  ERROR: {e}")
